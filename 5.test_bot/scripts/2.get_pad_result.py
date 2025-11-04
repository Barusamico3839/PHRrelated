# -*- coding: utf-8 -*-
import os
import sys
import time
from typing import Optional
import glob
import shutil
import re


def _import_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as e:
        print(f"[2.get_pad_result] openpyxl 未インストール: {e}")
        return None


def _import_uia() -> Optional[object]:
    try:
        import uiautomation as auto
        return auto
    except Exception as e:
        print(f"[2.get_pad_result] uiautomation 未インストールまたは利用不可: {e}")
        return None


def _ensure_results_book(dst_path: str):
    oxl = _import_openpyxl()
    if not oxl:
        raise RuntimeError("openpyxl が必要です")
    if os.path.exists(dst_path):
        return
    wb = oxl.Workbook()
    # 少なくとも1枚はシートを残す（空ブック保存対策）
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(dst_path)
    print(f"[2.get_pad_result] 作成 {dst_path}")


def _copy_sheet_values(src_path: str, src_sheet: str, dst_path: str, new_sheet_title: str) -> str:
    oxl = _import_openpyxl()
    if not oxl:
        raise RuntimeError("openpyxl が必要です")
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"ソースファイルが見つかりません: {src_path}")
    _ensure_results_book(dst_path)

    src_wb = oxl.load_workbook(src_path, data_only=True)
    if src_sheet not in src_wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {src_sheet}")
    s = src_wb[src_sheet]

    dst_wb = oxl.load_workbook(dst_path)
    # シートが1枚も無い場合に備える（念のため）
    if not dst_wb.sheetnames:
        dst_wb.create_sheet("Sheet1")
    title = str(new_sheet_title)
    base_title = title
    suffix = 1
    while title in dst_wb.sheetnames:
        suffix += 1
        title = f"{base_title}_{suffix}"
    d = dst_wb.create_sheet(title)

    for row in s.iter_rows():
        for cell in row:
            d.cell(row=cell.row, column=cell.column, value=cell.value)

    dst_wb.save(dst_path)
    print(f"[2.get_pad_result] '{src_sheet}' を '{dst_path}' のシート '{title}' にコピーしました")
    return title


def _import_win32():
    try:
        import win32com.client  # type: ignore
        return win32com.client
    except Exception as e:
        print(f"[2.get_pad_result] pywin32 インポート失敗: {e}")
        return None


def _scripts_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _ensure_results_book_template(dst_path: str):
    if os.path.exists(dst_path):
        return
    template_path = os.path.join(_scripts_dir(), "results_template.xlsx")
    try:
        if os.path.exists(template_path):
            shutil.copy2(template_path, dst_path)
            print(f"[2.get_pad_result] テンプレートから作成: {dst_path}")
            return
        else:
            print(f"[2.get_pad_result] テンプレートが見つかりません: {template_path}")
    except Exception as e:
        print(f"[2.get_pad_result] テンプレートコピー失敗のため空ブック作成にフォールバック: {e}")
    try:
        _ensure_results_book(dst_path)
    except Exception:
        pass


def _copy_sheet_full_via_com(src_path: str, src_sheet: str, dst_path: str, new_sheet_title: str):
    w32 = _import_win32()
    if not w32:
        return None
    excel = None
    src_wb = None
    dst_wb = None
    try:
        excel = w32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        src_wb = excel.Workbooks.Open(src_path)
        dst_wb = excel.Workbooks.Open(dst_path)
        try:
            ws = src_wb.Worksheets(src_sheet)
        except Exception:
            ws = src_wb.Worksheets(1)
        # Reliable cross-workbook copy using Before; will move to end later
        ws.Copy(Before=dst_wb.Worksheets(1))
        new_ws = excel.ActiveSheet
        base = str(new_sheet_title)
        name = base
        suffix = 1
        def _exists(nm: str) -> bool:
            try:
                _ = dst_wb.Worksheets(nm)
                return True
            except Exception:
                return False
        while _exists(name):
            suffix += 1
            name = f"{base}_{suffix}"
        try:
            new_ws.Name = name
        except Exception:
            name = f"Sheet_{int(time.time())%100000}"
            new_ws.Name = name
        # Move the new sheet to the end
        try:
            new_ws.Move(After=dst_wb.Worksheets(dst_wb.Worksheets.Count))
        except Exception:
            pass
        dst_wb.Save()
        print(f"[2.get_pad_result] 完全コピー: '{src_sheet}' -> '{dst_path}' の '{name}'")
        return name
    except Exception as e:
        print(f"[2.get_pad_result] Excel COM での完全コピー失敗: {e}")
        return None
    finally:
        try:
            if src_wb is not None:
                src_wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if dst_wb is not None:
                dst_wb.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


def _parse_emails(raw: str) -> list:
    if not raw:
        return []
    s = str(raw).replace("；", ";").replace("、", ";").replace(",", ";")
    parts = [p.strip() for p in s.split(";")]
    emails = []
    pat = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    for p in parts:
        if not p:
            continue
        m = pat.search(p)
        if m:
            emails.append(m.group(0))
    return emails


def _write_results_rows(dst_path: str, tehai_sheet_name: str, root_row: int, message_dialog: Optional[str]) -> None:
    oxl = _import_openpyxl()
    if not oxl:
        raise RuntimeError("openpyxl が必要です")
    wb = oxl.load_workbook(dst_path)
    if tehai_sheet_name not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {tehai_sheet_name}")
    if 'results' not in wb.sheetnames:
        wb.create_sheet('results')
    s = wb[tehai_sheet_name]
    r = wb['results']

    def val(addr: str):
        try:
            return s[addr].value
        except Exception:
            return None

    if message_dialog and (('エラー' in message_dialog) or ('error' in str(message_dialog).lower())):
        r[f"E{root_row + 1}"].value = str(message_dialog)

    mapping = [
        ("D3",  "E", 5),
        ("D12", "I", 5),
        ("D13", "M", 5),
        ("D15", "E", 8),
        ("D17", "I", 8),
        ("D18", "M", 8, True),
        ("D19", "E", 11, True),
        ("D20", "I", 11, True),
        ("D21", "M", 11, True),
    ]
    for item in mapping:
        src, col, off = item[0], item[1], item[2]
        skip_if_unused = (len(item) >= 4 and bool(item[3]))
        v = val(src)
        # For results M column at root_row+5, drop the first 5 characters
        if col == 'M' and off == 5:
            try:
                sv = str(v) if v is not None else ''
                v = sv[5:]
            except Exception:
                pass
        if skip_if_unused and (str(v) == '【使わない】'):
            continue
        r[f"{col}{root_row + off}"].value = v

    emails1 = _parse_emails(str(val('D25') or ''))
    emails2 = _parse_emails(str(val('D26') or ''))
    base_cols = ['E','F','G','H','I','J','K','L','M','N','O','P']
    for idx, addr in enumerate(emails1):
        if idx >= len(base_cols):
            break
        r[f"{base_cols[idx]}{root_row + 14}"].value = addr
    for idx, addr in enumerate(emails2):
        if idx >= len(base_cols):
            break
        r[f"{base_cols[idx]}{root_row + 17}"].value = addr

    wb.save(dst_path)
    print(f"[2.get_pad_result] results シートへ転記完了: root_row={root_row}")


def _try_click_buttons():
    auto = _import_uia()
    if not auto:
        print("[2.get_pad_result] UI操作スキップ")
        return

    def find_button_by(name=None, automation_id=None):
        try:
            root = auto.WindowControl()
            for c in root.GetChildren():
                try:
                    btn = None
                    if automation_id:
                        btn = c.ButtonControl(AutomationId=automation_id)
                        if btn and btn.Exists(0):
                            return btn
                    if name:
                        btn = c.ButtonControl(Name=name)
                        if btn and btn.Exists(0):
                            return btn
                except Exception:
                    continue
        except Exception:
            pass
        return None

    btn = find_button_by(name="はい", automation_id="3672912")
    if btn:
        try:
            btn.Click()
            print("[2.get_pad_result] 'はい' をクリック")
        except Exception as e:
            print(f"[2.get_pad_result] 'はい' クリック失敗: {e}")

    deadline = time.time() + 6
    while time.time() < deadline:
        ok = find_button_by(name="OK", automation_id="8590578")
        if ok:
            try:
                ok.Click()
                print("[2.get_pad_result] 'OK' をクリック")
                time.sleep(0.4)
                continue
            except Exception as e:
                print(f"[2.get_pad_result] 'OK' クリック失敗: {e}")
                break
        time.sleep(0.5)


def run(tehai_number: int, timestamp: str, week_index: int = 0, message_dialog: Optional[str] = None) -> str:
    print(f"[2.get_pad_result] 開始 tehai_number={tehai_number}, ts={timestamp}")
    src_path = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        "【全社標準】弔事対応フォルダ",
        "2. RPAブック",
        "RPAブック.xlsx",
    )
    dst_dir = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        "【全社標準】弔事対応フォルダ",
        "5.test_bot",
    )
    os.makedirs(dst_dir, exist_ok=True)
    dst_path = os.path.join(dst_dir, f"results_{timestamp}.xlsx")

    # テンプレートから結果ブックを用意し、RPAシートを完全コピー→resultsへ転記
    _ensure_results_book_template(dst_path)
    src_sheet_name = "RPAシート"
    new_sheet = _copy_sheet_full_via_com(src_path, src_sheet_name, dst_path, str(tehai_number))
    if not new_sheet:
        try:
            pass
        except Exception as e:
            print(f"[2.get_pad_result] Excelコピーエラー: {e}")
            raise
    try:
        root_row = 10 + (max(0, int(week_index)) * 20)
        _write_results_rows(dst_path, new_sheet, root_row, message_dialog)
    except Exception as e:
        print(f"[2.get_pad_result] results への転記エラー: {e}")

    try:
        pass
    except Exception as e:
        print(f"[2.get_pad_result] Excel処理エラー: {e}")
        raise

    # 添付PDFのコピー: '*弔事連絡票.pdf' を探し、results_%ts%添付PDF/{tehai_number}.pdf へ保存
    try:
        src_pdf_dir = os.path.join(
            os.path.expanduser("~"),
            "Desktop",
            "【全社標準】弔事対応フォルダ",
            "2. RPAブック",
        )
        patterns = [
            os.path.join(src_pdf_dir, "*弔事連絡票.pdf"),
            os.path.join(src_pdf_dir, "* 弔事連絡票.pdf"),
            os.path.join(src_pdf_dir, "*　弔事連絡票.pdf"),
        ]
        candidates = []
        # 追加: 正しい日本語名のパターンも含める
        try:
            patterns.extend([
                os.path.join(src_pdf_dir, "*弔事連絡票.pdf"),
                os.path.join(src_pdf_dir, "* 弔事連絡票.pdf"),
                os.path.join(src_pdf_dir, "*　弔事連絡票.pdf"),
            ])
        except Exception:
            pass
        for pat in patterns:
            try:
                candidates.extend(glob.glob(pat))
            except Exception:
                pass
        # 重複除去
        candidates = list(dict.fromkeys(candidates))
        if candidates:
            latest = max(candidates, key=lambda p: os.path.getmtime(p))
            pdf_dir = os.path.join(dst_dir, f"results_{timestamp}添付PDF")
            os.makedirs(pdf_dir, exist_ok=True)
            pdf_dst = os.path.join(pdf_dir, f"{tehai_number}.pdf")
            shutil.copy2(latest, pdf_dst)
            print(f"[2.get_pad_result] PDFコピー: '{latest}' -> '{pdf_dst}'")
        else:
            print(f"[2.get_pad_result] 弔事連絡票のPDFが見つかりませんでした: dir={src_pdf_dir}")
    except Exception as e:
        print(f"[2.get_pad_result] PDFコピー時のエラー: {e}")
    _try_click_buttons()
    print("[2.get_pad_result] 完了")
    return new_sheet


if __name__ == "__main__":
    try:
        run(1234, "0101_1234", 0, None)
    except Exception as e:
        print(f"[2.get_pad_result] 例外: {e}")
        sys.exit(1)








