# -*- coding: utf-8 -*-
import os
import sys
import re
from typing import Optional, List
from difflib import SequenceMatcher


def _import_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as e:
        print(f"[3.get_mail_result] openpyxl インポート失敗: {e}")
        return None


def _import_win32():
    try:
        import win32com.client  # type: ignore
        return win32com.client
    except Exception as e:
        print(f"[3.get_mail_result] pywin32 インポート失敗: {e}")
        return None


def _best_mail_match(sent_items, keyword: str, target: str):
    best = None
    best_score = -1.0
    count_checked = 0
    for item in sent_items:
        try:
            subj = str(getattr(item, 'Subject', '') or '')
            body = str(getattr(item, 'Body', '') or '')
            if keyword not in subj and keyword not in body:
                continue
            s1 = SequenceMatcher(None, subj, target).ratio()
            s2 = SequenceMatcher(None, body, target).ratio()
            score = max(s1, s2)
            if score > best_score:
                best = item
                best_score = score
            count_checked += 1
        except Exception:
            continue
    print(f"[3.get_mail_result] 走査メール数={count_checked}, ベストスコア={best_score:.3f}")
    return best


def _smtp_from_recipient(rec) -> Optional[str]:
    try:
        ae = getattr(rec, 'AddressEntry', None)
        if ae is not None:
            try:
                exu = ae.GetExchangeUser()
                if exu is not None:
                    addr = getattr(exu, 'PrimarySmtpAddress', None)
                    if addr:
                        return str(addr)
            except Exception:
                pass
            try:
                exd = ae.GetExchangeDistributionList()
                if exd is not None:
                    addr = getattr(exd, 'PrimarySmtpAddress', None)
                    if addr:
                        return str(addr)
            except Exception:
                pass
        a = getattr(rec, 'Address', None)
        if a:
            return str(a)
    except Exception:
        pass
    return None


def _recipients(mail, kind: int) -> List[str]:
    lst: List[str] = []
    try:
        for r in getattr(mail, 'Recipients', []):
            try:
                if getattr(r, 'Type', 0) == kind:
                    addr = _smtp_from_recipient(r)
                    if addr and addr not in lst:
                        lst.append(addr)
            except Exception:
                continue
    except Exception:
        pass
    return lst


def _parse_mail_body(body: str):
    lines = [ln.strip() for ln in (body or '').splitlines()]
    # 宛先人・本人氏名（〜様）
    name_lines = [ln for ln in lines if ln.endswith('様') and ln]
    addressee = None
    person = None
    if name_lines:
        addressee = name_lines[0].split('様')[0].strip()
        if len(name_lines) >= 2:
            person = name_lines[1].split('様')[0].strip()
    # 組合連絡
    union_status = None
    for ln in lines:
        m = re.search(r'組合連絡\s*：\s*(.+)$', ln)
        if m:
            union_status = m.group(1).strip()
            break
    # 弔電（先頭3文字）
    telegram_status = None
    for ln in lines:
        if ln.startswith('弔電'):
            part = ln.split('：', 1)[-1] if '：' in ln else ''
            part = part.strip().replace('　', ' ')
            telegram_status = part[:3]
            break
    # 供花（先頭3文字）
    flowers_status = None
    for ln in lines:
        if ln.startswith('供花'):
            part = ln.split('：', 1)[-1] if '：' in ln else ''
            part = part.strip().replace('　', ' ')
            flowers_status = part[:3]
            break
    # お香典（末尾2文字）
    condolence_status = None
    for ln in lines:
        if ln.startswith('お香典'):
            part = ln.split('：', 1)[-1] if '：' in ln else ''
            s = re.sub(r'\s+', '', part)
            condolence_status = s[-2:] if len(s) >= 2 else s
            break
    return addressee, person, union_status, telegram_status, flowers_status, condolence_status


def run(tehai_number: int, timestamp: str, sheet_name: Optional[str] = None, week_index: int = 0) -> None:
    print(f"[3.get_mail_result] 開始 tehai_number={tehai_number}, ts={timestamp}, week_index={week_index}")
    oxl = _import_openpyxl()
    if not oxl:
        raise RuntimeError("openpyxl が必要です")
    w32 = _import_win32()
    if not w32:
        print("[3.get_mail_result] Outlook連携をスキップします (pywin32 なし)")
        return

    dst_dir = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        "【全社標準】弔事対応フォルダ",
        "5.test_bot",
    )
    dst_path = os.path.join(dst_dir, f"results_{timestamp}.xlsx")
    if not os.path.exists(dst_path):
        raise FileNotFoundError(f"結果ブックが見つかりません: {dst_path}")

    wb = oxl.load_workbook(dst_path)
    target_sheet_name = sheet_name or str(tehai_number)
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {target_sheet_name}")
    ws = wb[target_sheet_name]

    d15 = str(ws["D15"].value or "")
    d27 = str(ws["D27"].value or "")
    if not d15:
        raise ValueError("D15 が空です (検索キーワード)")
    if not d27:
        print("[3.get_mail_result] D27 が空です。キーワードのみで検索します")

    outlook = w32.Dispatch("Outlook.Application")
    mapi = outlook.GetNamespace("MAPI")
    sent = mapi.GetDefaultFolder(5)  # 送信済みアイテム
    items = sent.Items
    try:
        items.Sort("[SentOn]", True)
    except Exception:
        pass

    mail = _best_mail_match(items, d15, d27 or d15)
    if not mail:
        raise RuntimeError("条件に一致する送信済みメールが見つかりませんでした")

    # 本文解析とヘッダ情報抽出
    content = str(getattr(mail, 'Body', '') or '')
    addressee, person, union_status, telegram_status, flowers_status, condolence_status = _parse_mail_body(content)
    subject = str(getattr(mail, 'Subject', '') or '')
    to_emails = _recipients(mail, 1)
    cc_emails = _recipients(mail, 2)
    bcc_emails = _recipients(mail, 3)
    att_names: List[str] = []
    try:
        atts = getattr(mail, 'Attachments', None)
        if atts:
            for i in range(1, int(getattr(atts, 'Count', 0)) + 1):
                try:
                    att = atts.Item(i)
                    nm = str(getattr(att, 'FileName', '') or '')
                    if nm:
                        att_names.append(nm)
                except Exception:
                    continue
    except Exception:
        pass

    # results シートへ書き込み
    r = wb['results'] if 'results' in wb.sheetnames else wb.create_sheet('results')
    root_row = 10 + (max(0, int(week_index)) * 20)
    # 宛先のe-mail/件名/添付名
    if to_emails:
        r[f"E{root_row + 6}"].value = to_emails[0]
    r[f"I{root_row + 6}"].value = subject
    s_att = "; ".join(att_names) if att_names else ""
    # For results M at root_row+6: drop first 5 characters and last 4 characters
    try:
        s_att = s_att[5:]
        s_att = s_att[:-4]
    except Exception:
        pass
    r[f"M{root_row + 6}"].value = s_att
    # 宛先人/本人氏名
    if addressee:
        r[f"E{root_row + 9}"].value = addressee
    if person:
        r[f"I{root_row + 9}"].value = person
    # 条件付きコピー
    try:
        m8 = str(r[f"M{root_row + 8}"].value or '')
        if union_status and (union_status in m8):
            r[f"M{root_row + 9}"].value = m8
    except Exception:
        pass
    try:
        e11 = str(r[f"E{root_row + 11}"].value or '')
        if telegram_status and (telegram_status in e11):
            r[f"E{root_row + 12}"].value = e11
    except Exception:
        pass
    try:
        i11 = str(r[f"I{root_row + 11}"].value or '')
        if flowers_status and (flowers_status in i11):
            r[f"I{root_row + 12}"].value = i11
    except Exception:
        pass
    try:
        m11 = str(r[f"M{root_row + 11}"].value or '')
        if condolence_status and (condolence_status in m11):
            r[f"M{root_row + 12}"].value = m11
    except Exception:
        pass

    # CC/BCC を 1行下にアラインして記入
    base_cols = ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R']
    def _row_emails_map(row_idx: int):
        mp = {}
        for c in base_cols:
            val = str(r[f"{c}{row_idx}"].value or '')
            if val:
                mp[val] = c
        return mp
    def _write_horiz(row_idx: int, emails: List[str], above_row: int):
        above = _row_emails_map(above_row)
        used = set()
        for addr in emails:
            col = above.get(addr)
            if col:
                r[f"{col}{row_idx}"].value = addr
                used.add(col)
        nxt_cols = [c for c in base_cols if c not in used]
        ci = 0
        for addr in emails:
            if addr in above:
                continue
            if ci >= len(nxt_cols):
                break
            r[f"{nxt_cols[ci]}{row_idx}"].value = addr
            ci += 1

    _write_horiz(root_row + 15, cc_emails, root_row + 14)
    _write_horiz(root_row + 18, bcc_emails, root_row + 17)

    wb.save(dst_path)
    print("[3.get_mail_result] メール解析→resultsへ反映 完了")


if __name__ == "__main__":
    try:
        run(1234, "0101_1234", None, 0)
    except Exception as e:
        print(f"[3.get_mail_result] エラー: {e}")
        sys.exit(1)
