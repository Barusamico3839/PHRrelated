#!/usr/bin/env python3
"""Append execution metadata to error_log.xlsx."""

from __future__ import annotations

import logging
from datetime import datetime

import tkinter as tk
from openpyxl import load_workbook  # type: ignore


LOGGER = logging.getLogger("chouji_robo.email")


def _ask_feedback(robot) -> str:
    window = tk.Toplevel(robot.root)
    window.title("フィードバック")
    window.geometry("560x320")
    window.attributes("-topmost", True)

    label = tk.Label(
        window,
        text="ロボに対してFB/コメント等があれば教えてください。（任意）",
        font=("游ゴシック", 14),
        anchor="w",
        justify="left",
    )
    label.pack(fill="x", padx=20, pady=(20, 8))

    text_box = tk.Text(window, width=70, height=8, font=("游ゴシック", 12))
    text_box.pack(padx=20, pady=4, fill="both", expand=True)

    result = {"value": ""}

    def _submit() -> None:
        result["value"] = text_box.get("1.0", "end").strip()
        window.destroy()

    button = tk.Button(window, text="確定", width=12, command=_submit)
    button.pack(pady=12)

    window.grab_set()
    window.wait_window(window)
    return result["value"]


def _first_empty_row(sheet) -> int:
    row = 2
    while True:
        if all((sheet.cell(row=row, column=col).value in (None, "")) for col in range(1, 9)):
            return row
        row += 1


def _format_duration(start: datetime, end: datetime) -> str:
    seconds = max(0, int((end - start).total_seconds()))
    minutes, sec = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return f"{hours:02}:{minutes:02}:{sec:02}"


def run(robot, *, success: bool, error_message: str) -> None:
    robot.current_phase = "Ed.error_log"
    workbook_path = robot.paths.error_log_book
    if not workbook_path.exists():
        LOGGER.warning("error_log.xlsx が見つからないため記録をスキップします: %s", workbook_path)
        return

    start_time = robot.state.workflow_started_at or datetime.now()
    end_time = robot.state.workflow_finished_at or datetime.now()
    duration_text = _format_duration(start_time, end_time)

    wb = load_workbook(workbook_path)
    sheet = wb.active

    target_row = _first_empty_row(sheet)

    sheet.cell(row=target_row, column=1).value = start_time
    sheet.cell(row=target_row, column=2).value = end_time
    sheet.cell(row=target_row, column=3).value = duration_text
    sheet.cell(row=target_row, column=4).value = robot.state.tehai_number or ""

    machine_id = robot._machine_identifier()
    sheet.cell(row=target_row, column=5).value = machine_id
    sheet.cell(row=target_row, column=6).value = "〇" if success else "✕"
    sheet.cell(row=target_row, column=7).value = error_message if not success else ""

    feedback = ""
    if success:
        feedback = _ask_feedback(robot)
    sheet.cell(row=target_row, column=8).value = feedback
    robot.state.workflow_feedback = feedback

    wb.save(workbook_path)
    LOGGER.info("error_log.xlsx に実行結果を記録しました (row=%s)。", target_row)
